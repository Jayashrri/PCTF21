from openpyxl import load_workbook
from openpyxl.styles import Alignment

r = ""

wb = load_workbook("./teachers.xlsx")
ws = wb['Worksheet']
for i in range(0, 231):
    cell = ws[f'A{i+2}']
    r += str(cell.alignment.text_rotation)

res = ""
for i in range(0,len(r),7):
    res += chr(int(r[i:i+7],2))

print(res)
